from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import F
from clubs.models import Club
from users.models import Membership
from .forms import SelfNominationForm, VotingForm
from .models import Election, Candidate, Vote, Position
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.text import slugify
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from io import BytesIO

@login_required
def active_elections(request):
    elections = Election.objects.filter(
        is_active=True,
        start_date__lte=timezone.now(),
        end_date__gte=timezone.now()
    )

    elections_with_status = []
    for election in elections:
        has_voted = Vote.objects.filter(election=election, voter=request.user).exists()
        elections_with_status.append({
            "election": election,
            "has_voted": has_voted
        })

    return render(request, "voting/elections.html", {"elections_with_status": elections_with_status})

@login_required
def cast_vote(request, election_id):
    election = get_object_or_404(Election, id=election_id, is_active=True)

    if timezone.now() < election.start_date or timezone.now() > election.end_date:
        messages.error(request, "Voting is not currently open for this election.")
        return redirect("voting:already_voted")

    # Use membership to check if user is a club member
    if not Membership.objects.filter(user=request.user, membership_type="member").exists():
        messages.error(request, "Only club members can vote.")
        return redirect("voting:already_voted")

    # Get clubs where user is a confirmed member
    user_clubs = Membership.objects.filter(
        user=request.user,
        membership_type="member"
    ).values_list("club_id", flat=True)

    # Get positions for which the user can vote
    positions = Position.objects.filter(election=election, club_id__in=user_clubs)

    # Get positions the user has already voted for
    voted_position_ids = Vote.objects.filter(
        election=election,
        voter=request.user,
        position__in=positions
    ).values_list("position_id", flat=True)

    if set(voted_position_ids) == set(positions.values_list("id", flat=True)) and positions.exists():
        messages.info(request, "You have already voted in this election.")
        return redirect("voting:already_voted")

    form = VotingForm(request.POST or None, election=election, user=request.user)

    if request.method == "POST":
        if form.is_valid():
            # Ensure every position has a selected candidate
            if not all(field in form.cleaned_data and form.cleaned_data[field] for field in form.fields):
                messages.warning(request, "You must vote for all available positions.")
                return redirect("voting:cast_vote", election_id=election.id)

            # First, remove any existing votes for these positions
            Vote.objects.filter(
                election=election,
                voter=request.user,
                position__in=positions
            ).delete()

            # Then create new votes
            for field_name, candidate in form.cleaned_data.items():
                Vote.objects.create(
                    election=election,
                    voter=request.user,
                    position=candidate.position,
                    candidate=candidate
                )

            messages.success(request, "Your votes have been recorded!")
            return redirect("voting:elections")
        else:
            messages.error(request, "There was a problem with your vote. Please try again.")

    return render(request, "voting/cast_vote.html", {
        "form": form,
        "election": election
    })

@login_required
@user_passes_test(lambda u: u.get_role() == "activity_center_admin")
def verify_results(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    election.results_verified = True
    election.save()

    assign_club_leaders(election)
    messages.success(request, "Results verified and club leaders updated.")
    return redirect("voting:results", election_id=election.id)

@login_required
def candidate_list(request, election_id, position_id):
    election = get_object_or_404(Election, id=election_id, is_active=True)
    position = get_object_or_404(Position, id=position_id, election=election)
    candidates = Candidate.objects.filter(election=election, position=position, approved=True)
    return render(request, "voting/candidates.html", {
        "election": election,
        "position": position,
        "candidates": candidates
    })

@login_required
def already_voted(request):
    return render(request, 'voting/already_voted.html')

@login_required
def thank_you(request):
    return render(request, 'voting/thank_you.html')

@login_required
def election_results(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    if timezone.now() < election.end_date or not election.results_verified:
        messages.warning(request, "Election results are not yet available.")
        return redirect("voting:elections")

    candidates = Candidate.objects.filter(election=election).order_by('-votes')
    return render(request, "voting/results.html", {
        "election": election,
        "candidates": candidates
    })

@login_required
def self_nominate(request):
    is_already_leader = Membership.objects.filter(
        user=request.user,
        membership_type="leader"
    ).exists()
    if is_already_leader:
        messages.error(request, "You are already a leader of a club and cannot run for another.")
        return redirect("club_member:dashboard")

    member_clubs = Club.objects.filter(
        memberships__user=request.user,
        memberships__membership_type="member"
    )
    if not member_clubs.exists():
        messages.error(request, "You are not a member of any club.")
        return redirect("club_member:dashboard")

    election = Election.get_current_election()
    if not election or not election.is_nomination_open():
        messages.error(request, "Nominations are currently closed.")
        return redirect("club_member:dashboard")

    positions_qs = Position.objects.filter(
        election=election,
        club__in=member_clubs
    )
    if not positions_qs.exists():
        messages.error(request, "No positions available to nominate for.")
        return redirect("club_member:dashboard")

    if request.method == "POST":
        form = SelfNominationForm(request.POST)
        form.fields["position"].queryset = positions_qs
        if form.is_valid():
            candidate = form.save(commit=False)
            # Prevent duplicate self-nomination
            if Candidate.objects.filter(user=request.user, election=election, position=candidate.position).exists():
                messages.warning(request, "You have already nominated yourself for this position.")
                return redirect("club_member:dashboard")

            candidate.user = request.user
            candidate.election = election
            candidate.club = candidate.position.club
            candidate.self_nominated = True
            candidate.approved = False
            candidate.save()
            messages.success(request, "Your nomination has been submitted.")
            return redirect("club_member:dashboard")
    else:
        form = SelfNominationForm()
        form.fields["position"].queryset = positions_qs

    return render(request, "voting/self_nomination.html", {"form": form})

@login_required
@user_passes_test(lambda u: u.is_staff or u.get_role() == "activity_center_admin")
def live_results(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    positions = Position.objects.filter(election=election).select_related("club")

    results = []
    for pos in positions:
        candidates = Candidate.objects.filter(
            election=election,
            position=pos,
            approved=True
        ).order_by("-votes")
        results.append({"position": pos, "candidates": candidates})

    return render(request, "voting/live_results.html", {
        "election": election,
        "results": results
    })

def assign_club_leaders(election):
    leader_positions = Position.objects.filter(election=election, name__iexact="President")
    for position in leader_positions:
        club = position.club
        top_candidate = Candidate.objects.filter(
            election=election,
            position=position,
            approved=True,
            club=club
        ).order_by("-votes").first()

        if top_candidate:
            # Demote existing leaders in this club
            Membership.objects.filter(club=club, membership_type="leader").exclude(user=top_candidate.user).update(membership_type="member")

            # Promote or create leader membership for the winner
            Membership.objects.update_or_create(
                user=top_candidate.user,
                club=club,
                defaults={"membership_type": "leader"}
            )


@login_required
@user_passes_test(lambda u: u.is_staff or u.get_role() == "activity_center_admin")
def export_results_excel(request, election_id):
    election = get_object_or_404(Election, id=election_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Election Results"

    # Header row
    ws.append(["Position", "Candidate", "Votes", "Club"])

    for position in Position.objects.filter(election=election):
        candidates = Candidate.objects.filter(election=election, position=position).order_by("-votes")
        for candidate in candidates:
            ws.append([
                position.name,
                candidate.user.name,
                candidate.votes,
                candidate.club.name if candidate.club else "—"
            ])

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"{slugify(election.name)}-results.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@user_passes_test(lambda u: u.is_staff or u.get_role() == "activity_center_admin")
def export_results_pdf(request, election_id):
    election = get_object_or_404(Election, id=election_id)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - inch
    p.setFont("Helvetica-Bold", 16)
    p.drawString(inch, y, f"Election Results - {election.name}")
    y -= 30

    p.setFont("Helvetica", 12)

    for position in Position.objects.filter(election=election):
        p.setFont("Helvetica-Bold", 13)
        p.drawString(inch, y, f"Position: {position.name} ({position.club.name if position.club else '—'})")
        y -= 20

        candidates = Candidate.objects.filter(election=election, position=position).order_by("-votes")
        if not candidates.exists():
            p.drawString(inch + 20, y, "No candidates for this position.")
            y -= 20
        else:
            p.setFont("Helvetica", 12)
            for candidate in candidates:
                text = f"- {candidate.user.name}: {candidate.votes} votes"
                p.drawString(inch + 20, y, text)
                y -= 18
                if y < inch:
                    p.showPage()
                    y = height - inch
                    p.setFont("Helvetica", 12)

        y -= 15

    p.showPage()
    p.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{election.name}-results.pdf"'
    return response
